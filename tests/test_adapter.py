# -*- coding: utf-8 -*-
"""
@File     :   test_adapter.py
@Desc     :   适配器层测试：消息协议 / 命令路由 / 世界管理 / 回合管线 / 回档 / Worker 挂接
@Note     :   全程 fake LLM 零网络；conftest 每测试独立临时库 + FakeLLM 注入；
             adapter 命令路由基于 storage/memory/worker 直连，不依赖真实网络
"""

from __future__ import annotations

import asyncio

import pytest

from src.adapter.base import AbstractAdapter
from src.adapter.web.adapter import WebAdapter
from src.adapter.protocol import InboundMessage, MessageType, OutboundMessage
from src.core.ids import make_world_id
from src.memory.fake import FakeMemory
from src.memory.interface import MemoryHit
from src.memory.worker import ConsolidationWorker

# 与 conftest 中临时模组目录预置的文件名保持一致（可被绑定校验通过）
TEST_MODULE_NAME = "test_module.docx"


# ============================================
# 消息协议
# ============================================


def test_protocol_factories():
    """入站/出站工厂：player_input / system_cmd / narrative / system_msg 类型与字段正确。"""
    pi = InboundMessage.player_input("调查员搜查房间", session_id="cli-default", world_id="world_001_x")
    assert pi.type == MessageType.PLAYER_INPUT
    assert pi.text == "调查员搜查房间"
    assert pi.world_id == "world_001_x"

    sc = InboundMessage.system_cmd("/status", session_id="cli-default")
    assert sc.type == MessageType.SYSTEM_CMD
    assert sc.text == "/status"

    n = OutboundMessage.narrative("叙事文本", session_id="cli-default")
    assert n.type == MessageType.NARRATIVE
    assert n.text == "叙事文本"

    s = OutboundMessage.system_msg("提示", level="warn", session_id="cli-default")
    assert s.type == MessageType.SYSTEM_MSG
    assert s.data["level"] == "warn"


# ============================================
# parse / 命令路由
# ============================================


async def test_parse_classifies_command_and_input(storage, world_id, fake_llm):
    """parse：/ 开头归 SYSTEM_CMD，其余归 PLAYER_INPUT，且映射会话当前世界。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    adapter._world_id = world_id

    cmd = await adapter.parse("/status")
    assert cmd.type == MessageType.SYSTEM_CMD
    assert cmd.world_id == world_id

    action = await adapter.parse("调查员搜查房间")
    assert action.type == MessageType.PLAYER_INPUT
    assert action.text == "调查员搜查房间"
    assert action.world_id == world_id


async def test_handle_unknown_command(storage, fake_llm):
    """未知命令返回 warn 提示，不抛异常。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/nope", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"


async def test_player_input_without_world(storage, fake_llm):
    """未选择世界时玩家输入被拦截并提示，不触发管线。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.player_input("调查员搜查房间", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert "未选择世界" in out.text


# ============================================
# 世界管理命令
# ============================================


async def test_world_start_creates_and_switches(storage, fake_llm):
    """/world start <世界id>：登记流程 -> 交互输入模组名 -> 跳过角色卡 -> 创建并切换。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/world start test1", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "世界 id 已登记" in out.text
    assert "模组名" in out.text

    out = await adapter.handle(
        InboundMessage.player_input(TEST_MODULE_NAME, session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "模组已确认" in out.text
    assert "角色卡名" in out.text

    out = await adapter.handle(
        InboundMessage.player_input("跳过", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "世界已创建" in out.text
    assert adapter._world_id == "test1"
    world = storage.get_world("test1")
    assert world is not None
    assert world["module_name"] == TEST_MODULE_NAME


async def test_world_start_requires_world_id(storage, fake_llm):
    """/world start 不带世界 id 时给出用法提示，不登记流程也不创建世界。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/world start", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"
    assert "用法" in out.text
    assert adapter._pending_world_flow.get("s") is None
    assert storage.list_worlds() == []


async def test_world_start_bad_module_rejected(storage, fake_llm):
    """交互中模组名不存在：提示重新输入并保持流程，不创建世界。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/world start test1", session_id="s")
    )
    assert "世界 id 已登记" in out.text

    out = await adapter.handle(
        InboundMessage.player_input("missing_module.pdf", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"
    assert "模组不存在" in out.text
    assert storage.list_worlds() == []

    # 状态：流程仍保持 module 步，可继续输入正确模组名
    out = await adapter.handle(
        InboundMessage.player_input(TEST_MODULE_NAME, session_id="s")
    )
    assert "模组已确认" in out.text


async def test_world_start_duplicate_world_id_rejected(storage, world_id, fake_llm):
    """世界 id 不允许重复：已存在的 id 直接拒绝，不进入交互流程。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd(f"/world start {world_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"
    assert "已存在" in out.text
    assert adapter._pending_world_flow.get("s") is None


async def test_world_start_invalid_world_id_rejected(storage, fake_llm):
    """非法世界 id（含空白/非法字符）拒绝，不登记流程。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/world start bad id", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"
    assert adapter._pending_world_flow.get("s") is None

    out = await adapter.handle(
        InboundMessage.system_cmd("/world start bad/id", session_id="s")
    )
    assert "只能包含" in out.text
    assert adapter._pending_world_flow.get("s") is None


async def test_world_list_and_load(storage, world_id, fake_llm):
    """/world list 列出世界；/world load 载入当前世界。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    listed = await adapter.handle(InboundMessage.system_cmd("/world list", session_id="s"))
    assert listed.type == MessageType.SYSTEM_MSG
    assert world_id in listed.text

    loaded = await adapter.handle(InboundMessage.system_cmd(f"/world load {world_id}", session_id="s"))
    assert "已载入" in loaded.text
    assert adapter._world_id == world_id


async def test_world_load_unknown(storage, fake_llm):
    """/world load 不存在的世界返回 warn。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/world load world_999_nope", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"


async def test_world_load_empty_recall(storage, world_id, fake_llm):
    """/world load 载入无历史世界：提示暂无历史记录，不抛异常。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd(f"/world load {world_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "已载入世界" in out.text
    assert "暂无历史记录" in out.text
    assert adapter._world_id == world_id


async def test_world_load_shows_last_reply(storage, world_id, fake_llm):
    """/world load 显示最新程序回复（最近一轮的玩家视角叙事），帮玩家接续存档。"""
    storage.commit_turn(
        world_id, 1,
        context_data={
            "user": "调查员叩响门环",
            "assistant": "门内传来脚步声，一位老人打开了门。",
        },
    )
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd(f"/world load {world_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "【上次进展】" in out.text
    assert "门内传来脚步声，一位老人打开了门。" in out.text


async def test_world_load_shows_memories(storage, world_id, fake_llm):
    """/world load 有记忆后端时显示最近记忆条目（· 项目符号排版）。"""
    class _StubMemory:
        def __init__(self, hits):
            self.hits = hits

        async def search(self, query, world_id, *, top_k=5, since_turn=None):
            return self.hits

    mem = _StubMemory(
        [MemoryHit(text="调查员在公墓发现墓碑上的读书人影", turn_num=2, score=0.9)]
    )
    adapter = WebAdapter(storage=storage, memory=mem, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd(f"/world load {world_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "【近期记忆】" in out.text
    assert "· 调查员在公墓发现墓碑上的读书人影" in out.text
    assert "[t" not in out.text


async def test_world_delete_removes_world(storage, world_id, fake_llm):
    """/world delete 删除世界并取消会话选中（SQLite 级联 + RAG 清理）。"""
    from src.memory.fake import FakeMemory

    storage.create_entity(
        world_id, "pc_01", "PC", "费莉西蒂",
        hp=8, hp_max=12, san=58, san_max=70,
        attributes_and_skills={"侦查": 60},
    )
    memory = FakeMemory(storage=storage)
    adapter = WebAdapter(storage=storage, memory=memory, llm=fake_llm.call)
    adapter._world_id = world_id

    out = await adapter.handle(
        InboundMessage.system_cmd(f"/world delete {world_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "世界已删除" in out.text
    assert storage.get_world(world_id) is None
    assert storage.get_entities(world_id) == []  # 状态：外键级联实体一并删除
    assert adapter._world_id is None  # 状态：删除当前选中世界后取消会话指向


async def test_world_delete_unknown(storage, fake_llm):
    """/world delete 不存在的世界返回 warn，不误删。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/world delete world_999_nope", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "warn"
    assert "世界不存在" in out.text


# ============================================
# 状态查询
# ============================================


async def test_status_without_world(storage, fake_llm):
    """/status 未选择世界时提示。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/status", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert "未选择世界" in out.text


async def test_status_shows_world_and_pc(storage, world_id, fake_llm):
    """/status 显示世界模组、最新轮次与 PC 实体。"""
    storage.create_entity(
        world_id, "pc_01", "PC", "费莉西蒂",
        hp=8, hp_max=12, san=58, san_max=70,
        attributes_and_skills={"侦查": 60},
    )
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    adapter._world_id = world_id
    out = await adapter.handle(InboundMessage.system_cmd("/status", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert "费莉西蒂" in out.text
    assert "HP 8/12" in out.text


# ============================================
# 回合管线
# ============================================


async def _step_directive(narrative="### 规则裁决\n- 侦查成功"):
    """首轮直接 present_directive 交卷的 fake 响应（工具名须与 schemas 一致）。"""

    def step(messages):
        return {"text": None, "tool_calls": [
            {"id": "c1", "name": "present_directive",
             "arguments": {"narrative_directive": narrative}}]}

    return step


async def _step_stats_then_directive(narrative="### 规则裁决\n- 侦查成功（18/60）"):
    """首轮 check_and_update_stats 扣血（含检定），回填后 present_directive 交卷，产生 hp diff。"""

    def step(messages):
        if any(m["role"] == "tool" for m in messages):
            return {"text": None, "tool_calls": [
                {"id": "c2", "name": "present_directive",
                 "arguments": {"narrative_directive": narrative}}]}
        return {"text": None, "tool_calls": [
            {"id": "c1", "name": "check_and_update_stats",
             "arguments": {"entity_id": "pc_01", "skill_or_attribute": "侦查", "hp_change": -3}}]}

    return step


async def test_player_input_runs_pipeline(storage, world_id, fake_llm):
    """玩家输入在有世界时触发回合管线，返回玩家视角叙事，并落库轮次。"""
    fake_llm.set_response("smart", await _step_directive())
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    adapter._world_id = world_id

    out = await adapter.handle(InboundMessage.player_input("调查员试图撬开暗门", session_id="s"))
    assert out.type == MessageType.NARRATIVE
    assert out.text  # 叙事非空
    assert storage.next_turn_num(world_id) == 2  # 状态：已落库 1 轮


async def test_pipeline_hook_triggers_worker(storage, world_id, fake_llm):
    """落库后 on_turn_committed 钩子触发 worker.trigger_world（事件触发通道接通）。"""
    fake_llm.set_response("smart", await _step_directive())
    fired = []

    class _WorkerStub:
        def trigger_world(self, wid, *, force=False):
            fired.append((wid, force))

    adapter = WebAdapter(storage=storage, llm=fake_llm.call, worker=_WorkerStub())
    adapter._world_id = world_id
    out = await adapter.handle(InboundMessage.player_input("调查员搜查房间", session_id="s"))
    assert out.type == MessageType.NARRATIVE
    assert fired == [(world_id, False)]


# ============================================
# 回档命令
# ============================================


async def test_rollback_requires_world(storage, fake_llm):
    """/rollback 未选择世界时提示。"""
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/rollback", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert "未选择世界" in out.text


async def test_rollback_lists_and_executes(storage, world_id, fake_llm):
    """/rollback 无参列出最近轮次；/rollback <N> 双侧回档（SQLite + RAG）。"""
    storage.create_entity(
        world_id, "pc_01", "PC", "费莉西蒂",
        hp=10, hp_max=12, san=58, san_max=70,
        attributes_and_skills={"侦查": 60},
    )
    fake = FakeMemory(storage=storage)  # 状态：假门面内部自带 FakeMemoryBackend
    fake_llm.set_response("smart", await _step_stats_then_directive())
    adapter = WebAdapter(storage=storage, memory=fake, llm=fake_llm.call)
    adapter._world_id = world_id

    hp_before = storage.get_entity(world_id, "pc_01")["hp"]  # 10
    # 跑一轮带扣血的回合，制造可回档的 hp diff
    await adapter.handle(InboundMessage.player_input("调查员撞破暗门", session_id="s"))
    assert storage.get_entity(world_id, "pc_01")["hp"] == hp_before - 3

    listed = await adapter.handle(InboundMessage.system_cmd("/rollback", session_id="s"))
    assert listed.type == MessageType.SYSTEM_MSG
    assert "最新轮次" in listed.text

    rolled = await adapter.handle(InboundMessage.system_cmd("/rollback 1", session_id="s"))
    assert rolled.type == MessageType.SYSTEM_MSG
    assert "已回档" in rolled.text
    # 物理状态回到回档前快照（本轮 hp 变更被撤销）
    assert storage.get_entity(world_id, "pc_01")["hp"] == hp_before
    assert storage.next_turn_num(world_id) == 1  # 状态：轮次已回退到空


# ============================================
# Worker 生命周期（runtime 编排）
# ============================================


async def test_worker_hook_binding(storage, world_id, fake_llm):
    """ConsolidationWorker 挂接：触发后进入待处理状态，不阻塞且失败不抛出。"""
    fake_llm.set_response("smart", await _step_directive())
    memory = FakeMemory(storage=storage)
    worker = ConsolidationWorker(memory=memory, storage=storage, worlds=[world_id])

    task = asyncio.create_task(worker.start())
    try:
        await asyncio.sleep(0)  # 状态：让 worker 进入等待循环
        adapter = WebAdapter(storage=storage, memory=memory, worker=worker, llm=fake_llm.call)
        adapter._world_id = world_id
        out = await adapter.handle(InboundMessage.player_input("调查员搜查房间", session_id="s"))
        assert out.type == MessageType.NARRATIVE
        assert worker.is_running  # 状态：事件触发不破坏 worker 生命周期
    finally:
        await worker.stop()
        if not task.done():
            task.cancel()
            try:
                await task
            except asyncio.CancelledError:
                pass


# ============================================
# 适配器工厂（create_adapter）
# ============================================


def _patch_adapter_settings(monkeypatch, data):
    """把 runtime 的 get_settings 替换为返回固定 adapter 段的假配置。"""
    from src.adapter import runtime as runtime_mod

    class _FakeSettings:
        def get(self, dotted_path, default=None):
            return data.get(dotted_path, default)

    monkeypatch.setattr(runtime_mod, "get_settings", lambda: _FakeSettings())


def test_create_adapter_web(monkeypatch):
    """config.adapter.active=web 时工厂分派 WebAdapter，并读取 web.session_id。"""
    _patch_adapter_settings(
        monkeypatch,
        {"adapter.active": "web", "adapter.web.session_id": "s-test"},
    )
    from src.adapter.runtime import create_adapter

    adapter = create_adapter(storage=None, memory=None, worker=None)
    assert isinstance(adapter, WebAdapter)
    assert adapter.session_id == "s-test"


def test_create_adapter_unknown(monkeypatch):
    """config.adapter.active 指向未注册类型时工厂抛 ValueError。"""
    _patch_adapter_settings(monkeypatch, {"adapter.active": "onebot"})
    from src.adapter.runtime import create_adapter

    with pytest.raises(ValueError):
        create_adapter(storage=None, memory=None, worker=None)


# ============================================
# 角色卡命令（/card）
# ============================================


def _make_min_card(path) -> None:
    """构造一张最小"人物卡" xlsx（骰子工厂布局坐标）。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "人物卡"
    ws.cell(row=3, column=5, value="费莉西蒂·利丝")   # E3 姓名
    ws.cell(row=6, column=13, value="女")             # M6 性别
    ws.cell(row=6, column=5, value=24)                # E6 年龄
    ws.cell(row=7, column=5, value="伦敦")            # E7 住地
    ws.cell(row=5, column=5, value="私家侦探")        # E5 职业
    stats = [
        ("STR", 21, 3, 50), ("DEX", 27, 3, 60), ("POW", 33, 3, 70),
        ("CON", 21, 5, 60), ("APP", 27, 5, 50), ("EDU", 33, 5, 70),
        ("SIZ", 21, 7, 50), ("INT", 27, 7, 70),
    ]
    for _, col, row, v in stats:
        ws.cell(row=row, column=col, value=v)
    wb.save(path)


async def test_card_list_empty(storage, tmp_path, monkeypatch, fake_llm):
    """种子库为空时 /card list 提示先导入。"""
    from src.tools import card_store

    monkeypatch.setattr(card_store, "SEED_DIR", tmp_path / "seeds")
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/card list", session_id="s"))
    assert out.type == MessageType.SYSTEM_MSG
    assert "种子库为空" in out.text


async def test_card_import_and_world_start_with_pc(storage, tmp_path, monkeypatch, fake_llm):
    """/card import 入种子库 + /world start 交互两步流程输入角色卡名建世界并绑定 PC。"""
    from src.tools import card_importer, card_store

    seeds = tmp_path / "seeds"
    cards = tmp_path / "cards"
    cards.mkdir()
    monkeypatch.setattr(card_store, "SEED_DIR", seeds)
    monkeypatch.setattr(card_importer, "CARDS_DIR", cards)
    _make_min_card(cards / "费莉西蒂.xlsx")

    adapter = WebAdapter(storage=storage, llm=fake_llm.call)

    out = await adapter.handle(
        InboundMessage.system_cmd("/card import 费莉西蒂.xlsx", session_id="s")
    )
    assert "角色卡已导入种子库" in out.text
    seed_rows = card_store.list_seed_cards()
    assert len(seed_rows) == 1
    assert "费莉西蒂" in seed_rows[0]["name"]
    assert seed_rows[0]["occupation"] == "私家侦探"

    out = await adapter.handle(
        InboundMessage.system_cmd("/world start test1", session_id="s")
    )
    assert "世界 id 已登记" in out.text
    out = await adapter.handle(
        InboundMessage.player_input(TEST_MODULE_NAME, session_id="s")
    )
    assert "模组已确认" in out.text
    out = await adapter.handle(
        InboundMessage.player_input("费莉西蒂", session_id="s")
    )
    assert "世界已创建并切换" in out.text
    assert "已绑定 PC" in out.text
    assert adapter._world_id == "test1"
    pcs = storage.get_entities("test1", entity_type="PC")
    assert len(pcs) == 1
    assert pcs[0]["name"].startswith("费莉西蒂")
    # 状态：模组 + PC 双要素齐备，自动顺承 Turn 0 开场演播（三件套落库）
    assert out.type == MessageType.NARRATIVE
    turn0 = storage.get_turn("test1", 0)
    assert turn0 is not None
    assert turn0["context_data"]["assistant"]
    assert turn0["solidified"] == 1  # 状态：已标记固化，后台 Worker 不会二次提炼开场


async def test_card_use_copies_to_current_world(storage, world_id, tmp_path, monkeypatch, fake_llm):
    """/card use 把种子角色拷贝到当前世界并绑定。"""
    from src.tools import card_store

    monkeypatch.setattr(card_store, "SEED_DIR", tmp_path / "seeds")
    seed_id = card_store.save_seed(
        {
            "entity_type": "PC", "name": "约翰", "hp": 10, "hp_max": 12,
            "mp": 9, "mp_max": 9, "san": 60, "san_max": 99,
            "attributes_and_skills": {}, "inventory": [], "background": {}, "tags": [],
        },
        {"name": "约翰", "occupation": "作家"}, source="x",
    )
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    adapter._world_id = world_id
    out = await adapter.handle(InboundMessage.system_cmd(f"/card use {seed_id}", session_id="s"))
    assert "已拷贝到当前世界" in out.text
    pcs = storage.get_entities(world_id, entity_type="PC")
    assert any(p["name"] == "约翰" for p in pcs)


async def test_card_use_without_world(storage, tmp_path, monkeypatch, fake_llm):
    """未选世界时 /card use 提示先建世界。"""
    from src.tools import card_store

    monkeypatch.setattr(card_store, "SEED_DIR", tmp_path / "seeds")
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(InboundMessage.system_cmd("/card use card_xxx", session_id="s"))
    assert "当前未选择世界" in out.text


async def test_card_delete_removes_seed(storage, tmp_path, monkeypatch, fake_llm):
    """/card delete 删除种子库中的角色卡文件。"""
    from src.tools import card_store

    monkeypatch.setattr(card_store, "SEED_DIR", tmp_path / "seeds")
    seed_id = card_store.save_seed(
        {
            "entity_type": "PC", "name": "约翰", "hp": 10, "hp_max": 12,
            "mp": 9, "mp_max": 9, "san": 60, "san_max": 99,
            "attributes_and_skills": {}, "inventory": [], "background": {}, "tags": [],
        },
        {"name": "约翰", "occupation": "作家"}, source="x",
    )
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd(f"/card delete {seed_id}", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert "种子卡已删除" in out.text
    assert card_store.list_seed_cards() == []


async def test_card_delete_unknown(storage, tmp_path, monkeypatch, fake_llm):
    """/card delete 不存在的种子返回 error。"""
    from src.tools import card_store

    monkeypatch.setattr(card_store, "SEED_DIR", tmp_path / "seeds")
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/card delete card_xxx", session_id="s")
    )
    assert out.type == MessageType.SYSTEM_MSG
    assert out.data["level"] == "error"
    assert "删除种子卡失败" in out.text


async def test_card_import_bad_source(storage, tmp_path, monkeypatch, fake_llm):
    """导入不存在的文件返回明确错误。"""
    from src.tools import card_importer

    monkeypatch.setattr(card_importer, "CARDS_DIR", tmp_path / "empty_cards")
    adapter = WebAdapter(storage=storage, llm=fake_llm.call)
    out = await adapter.handle(
        InboundMessage.system_cmd("/card import 不存在.xlsx", session_id="s")
    )
    assert out.data["level"] == "error"
    assert "角色卡导入失败" in out.text
