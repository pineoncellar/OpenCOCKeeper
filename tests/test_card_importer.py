# -*- coding: utf-8 -*-
"""
@File     :   test_card_importer.py
@Desc     :   角色卡 xlsx 解析测试：基本信息/八属性/技能合并/背景/背包/衍生值/入库
@Note     :   用 openpyxl 构造最小"人物卡" sheet 按 glyphkeeper 布局写入，再解析断言；
             产出 {meta, entity}，entity 可直接喂 storage.create_entity
"""

from __future__ import annotations

from pathlib import Path

from src.tools.card_importer import (
    _is_label,
    parse_investigator_xlsx,
)


def _make_card(path: Path, *, sheet: str = "人物卡") -> Path:
    """构造一张最小 xlsx 角色卡（骰子工厂布局坐标）。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    # 基本信息
    ws.cell(row=3, column=5, value="费莉西蒂·利丝")   # E3 姓名
    ws.cell(row=6, column=13, value="女")             # M6 性别
    ws.cell(row=6, column=5, value=24)                # E6 年龄
    ws.cell(row=7, column=5, value="伦敦")            # E7 住地
    ws.cell(row=5, column=5, value="私家侦探")        # E5 职业
    # 八属性
    stat_cells = {
        "STR": (21, 3), "DEX": (27, 3), "POW": (33, 3),
        "CON": (21, 5), "APP": (27, 5), "EDU": (33, 5),
        "SIZ": (21, 7), "INT": (27, 7),
    }
    stat_vals = {"STR": 50, "DEX": 60, "POW": 70, "CON": 60,
                 "APP": 50, "EDU": 70, "SIZ": 50, "INT": 70}
    for key, (col, row) in stat_cells.items():
        ws.cell(row=row, column=col, value=stat_vals[key])
    ws.cell(row=7, column=33, value=50)               # AG7 幸运
    # 技能：左半 F16 侦查（init+growth+occ+interest）、F17 图书馆使用→映射
    ws.cell(row=16, column=6, value="侦查")
    ws.cell(row=16, column=10, value=20)
    ws.cell(row=16, column=12, value=10)
    ws.cell(row=16, column=14, value=20)
    ws.cell(row=16, column=16, value=10)
    ws.cell(row=17, column=6, value="图书馆使用")
    ws.cell(row=17, column=10, value=5)
    # 右半 AB16 斗殴
    ws.cell(row=16, column=28, value="斗殴")
    ws.cell(row=16, column=32, value=25)
    # 背景（AA 列）
    ws.cell(row=61, column=27, value="略卷的蓝褐色长发")
    ws.cell(row=63, column=27, value="真正的力量")
    ws.cell(row=65, column=27, value="失踪的马瑟斯先生")
    ws.cell(row=67, column=27, value="马瑟斯的事务所")
    ws.cell(row=69, column=27, value="和塞恩的合影")
    ws.cell(row=71, column=27, value="言语温和")
    ws.cell(row=73, column=27, value="无")
    ws.cell(row=75, column=27, value="无")            # 恐惧症和躁狂症
    ws.cell(row=77, column=23, value="费莉西蒂·利丝的少女时代……")  # W77 完整背景故事
    # 背包
    ws.cell(row=78, column=6, value="左轮手枪")
    ws.cell(row=78, column=14, value="笔记本")
    wb.save(path)
    return path


def test_parse_meta(tmp_path):
    """meta：基本信息完整读取（含 luck）。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    data = parse_investigator_xlsx(card)
    meta = data["meta"]
    assert meta["name"] == "费莉西蒂·利丝"
    assert meta["gender"] == "女"
    assert meta["age"] == 24
    assert meta["birthplace"] == "伦敦"
    assert meta["occupation"] == "私家侦探"
    assert meta["luck"] == 50


def test_parse_entity_stats_and_skills(tmp_path):
    """entity：八属性进 attributes_and_skills，技能合并 init+growth+occ+interest 并映射。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    entity = parse_investigator_xlsx(card)["entity"]
    aas = entity["attributes_and_skills"]
    # 八属性
    assert aas["STR"] == 50 and aas["DEX"] == 60 and aas["POW"] == 70
    assert aas["CON"] == 60 and aas["APP"] == 50 and aas["EDU"] == 70
    assert aas["SIZ"] == 50 and aas["INT"] == 70
    # 技能合并与映射（"图书馆使用" → "图书馆利用"）
    assert aas["侦查"] == 20 + 10 + 20 + 10
    assert aas["图书馆利用"] == 5
    assert aas["斗殴"] == 25


def test_parse_entity_occupation(tmp_path):
    """entity：职业读角色卡 E5 格子。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    entity = parse_investigator_xlsx(card)["entity"]
    assert entity["occupation"] == "私家侦探"


def test_parse_entity_derived_hp_san_mp(tmp_path):
    """entity：HP/SAN/MP 由八属性衍生（CON+SIZ//10、POW//5、POW），满状态入场。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    entity = parse_investigator_xlsx(card)["entity"]
    assert entity["hp"] == entity["hp_max"] == (60 + 50) // 10  # 11
    assert entity["mp"] == entity["mp_max"] == max(1, 70 // 5)  # 14
    assert entity["san"] == entity["san_max"] == 70
    assert entity["entity_type"] == "PC"


def test_parse_entity_background_and_inventory(tmp_path):
    """entity：背景 9 字段（含完整背景故事/恐惧症）与背包。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    entity = parse_investigator_xlsx(card)["entity"]
    bg = entity["background"]
    assert bg["appearance_desc"] == "略卷的蓝褐色长发"
    assert bg["belief"] == "真正的力量"
    assert bg["significant_person"] == "失踪的马瑟斯先生"
    assert bg["significant_place"] == "马瑟斯的事务所"
    assert bg["cherished_possession"] == "和塞恩的合影"
    assert bg["trait"] == "言语温和"
    assert bg["injury_scar"] == "无"
    assert bg["phobias_manias"] == "无"
    assert bg["full_backstory"] == "费莉西蒂·利丝的少女时代……"
    assert entity["inventory"] == [{"name": "左轮手枪"}, {"name": "笔记本"}]
    assert entity["tags"] == []


def test_parse_entity_feed_create_entity(tmp_path, storage, world_id):
    """entity 可直接喂 storage.create_entity，落库后字段还原。"""
    card = _make_card(tmp_path / "felicity.xlsx")
    entity = parse_investigator_xlsx(card)["entity"]
    storage.create_entity(world_id, "pc_01", **entity)
    saved = storage.get_entity(world_id, "pc_01")
    assert saved["name"] == "费莉西蒂·利丝"
    assert saved["hp"] == 11 and saved["hp_max"] == 11
    assert saved["san"] == 70 and saved["san_max"] == 70
    assert saved["attributes_and_skills"]["侦查"] == 60
    assert saved["background"]["belief"] == "真正的力量"
    assert saved["inventory"] == [{"name": "左轮手枪"}, {"name": "笔记本"}]


def test_parse_empty_background(tmp_path):
    """背景全空时 background 为不含 key 的空 dict，不报错。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "人物卡"
    ws.cell(row=3, column=5, value="无名调查员")
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    entity = parse_investigator_xlsx(path)["entity"]
    assert entity["background"] == {}


def test_resolve_card_source_auto_extension(tmp_path, monkeypatch):
    """resolve_card_source：缺扩展名自动补 .xlsx（"费利西蒂"命中"费利西蒂.xlsx"）。"""
    import pytest

    from src.tools import card_importer

    d = tmp_path / "cards"
    d.mkdir()
    card = d / "费利西蒂.xlsx"
    card.write_bytes(b"x")
    monkeypatch.setattr(card_importer, "CARDS_DIR", d)
    # 缺扩展名自动补
    assert card_importer.resolve_card_source("费利西蒂") == card
    # 带扩展名精确命中
    assert card_importer.resolve_card_source("费利西蒂.xlsx") == card
    # 已有其他扩展名不补
    xls = d / "卡.xls"
    xls.write_bytes(b"x")
    assert card_importer.resolve_card_source("卡.xls") == xls
    # 0 命中：报错
    with pytest.raises(FileNotFoundError):
        card_importer.resolve_card_source("不存在的卡")


def test_is_label():
    """标签判定：多行/含标签关键词的单元格判为标签。"""
    assert _is_label("姓名")
    assert _is_label("力量\nSTR")
    assert not _is_label("费莉西蒂·利丝")
    assert not _is_label("")
