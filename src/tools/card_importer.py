# -*- coding: utf-8 -*-
"""
@File     :   card_importer.py
@Desc     :   从骰子工厂格式 Excel 角色卡（xlsx）读取调查员数据，产出新架构实体 dict
@Note     :   仅支持当前已知的 .xlsx 布局（坐标常量照搬 glyphkeeper/src/tools/card_importer.py）；
             只读取表格不写库；产出分 meta（原始元数据，供展示/选择）与 entity
             （可直接喂 storage.create_entity 的字段）；
             衍生值 HP/SAN/MP 由八属性自动计算（对齐 glyphkeeper：无视卡中数值，满状态入场）
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

from ..core.config import PROJECT_ROOT
from ..core.log import get_logger

logger = get_logger(__name__)


# ====================================================================
# 角色卡源文件目录
# ====================================================================

# 存放原始 .xlsx 角色卡文件的默认目录（自动创建），/card import 在此检索
CARDS_DIR = PROJECT_ROOT / "data" / "cards"


def _cards_dir() -> Path:
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    return CARDS_DIR


def list_cards() -> List[Path]:
    """列出 data/cards 目录下全部 xlsx 角色卡文件（按文件名排序）。"""
    if not CARDS_DIR.is_dir():
        return []
    return sorted(
        p for p in CARDS_DIR.iterdir()
        if p.suffix.lower() in (".xlsx", ".xls")
    )


def search_cards_dir(keyword: str) -> List[Path]:
    """在 data/cards 目录搜索匹配关键字的 xlsx 文件（忽略大小写与空格）。"""
    kw = (keyword or "").lower().replace(" ", "")
    return [p for p in list_cards() if kw in p.stem.lower().replace(" ", "")]


def resolve_card_source(source: str) -> Path:
    """解析 /card import 的来源：显式路径优先，其次为 cards 目录内文件名。

    找不到抛 FileNotFoundError（带可用列表提示）。
    """
    raw = (source or "").strip()
    if not raw:
        raise FileNotFoundError("用法: /card import <xlsx路径或 data/cards 内文件名>")
    path = Path(raw)
    if path.is_file():
        return path
    # 视为 cards 目录内文件名
    candidate = CARDS_DIR / raw
    if candidate.is_file():
        return candidate
    available = "、".join(p.stem for p in list_cards()) or "（空）"
    raise FileNotFoundError(
        f"角色卡文件不存在: {raw}\n当前 data/cards 可用: {available}"
    )


# ====================================================================
# 技能名映射
# ====================================================================

# 冲突名显式映射；编号技能（科学①、外语①、技艺①等）与 Ω 后缀名保留原名，
# 让运行时子串匹配兜底（对齐 glyphkeeper SKILL_NAME_MAP）
SKILL_NAME_MAP: dict[str, str] = {
    "图书馆使用": "图书馆利用",
    "母语": "语言(母语)",
    "驯兽": "动物驯养",
}

_STRIP_SUFFIX_RE = re.compile(r"[Ω：\u3000\s]+$")


def _normalize(raw: Any) -> str:
    """标准化技能名：去前后空白 → 映射 → 去 Ω/全角冒号后缀"""
    if not raw:
        return ""
    name = str(raw).strip()
    mapped = SKILL_NAME_MAP.get(name)
    if mapped:
        return mapped
    return _STRIP_SUFFIX_RE.sub("", name)


def _to_int(v: Any) -> int:
    """安全转 int，非数字返回 0"""
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (ValueError, AttributeError, TypeError):
        return 0


# ====================================================================
# 人物卡 Sheet 坐标常量（openpyxl 1-indexed）
# ====================================================================

_ROW_BASIC = 3        # 姓名、STR/DEX/POW
_ROW_JOB = 5          # 职业、CON/APP/EDU
_ROW_SIZ_INT = 7      # SIZ/INT/Luck
_ROW_SKILL_START = 16  # 技能表起始行
_ROW_SKILL_END = 54   # 技能表结束行
_ROW_INV_START = 78   # 背包起始
_ROW_INV_END = 91     # 背包结束
_COL_NAME = 5         # E 列
# 属性坐标：每行三个属性
_STAT_COLS = [
    ("STR", 21, 3),   # U3
    ("DEX", 27, 3),   # AA3
    ("POW", 33, 3),   # AG3
    ("CON", 21, 5),   # U5
    ("APP", 27, 5),   # AA5
    ("EDU", 33, 5),   # AG5
    ("SIZ", 21, 7),   # U7
    ("INT", 27, 7),   # AA7
]
_LUCK_COL = 33        # AG
_LUCK_ROW = 7

# 技能表：左半 (F/J/L/N/P) 和 右半 (AB/AF/AH/AJ/AL)
_SKILL_LEFT = dict(name=6, init=10, growth=12, occ=14, interest=16)
_SKILL_RIGHT = dict(name=28, init=32, growth=34, occ=36, interest=38)

# 背景：AA 列 (col=27) 存值 / W 列 (col=23) 存标签
_BG_COL = 27
_BG_ROWS: dict[str, int] = {
    "appearance_desc": 61,
    "belief": 63,
    "significant_person": 65,
    "significant_place": 67,
    "cherished_possession": 69,
    "trait": 71,
    "injury_scar": 73,
}
_PHOBIA_ROW = 75          # AA75 恐惧症和躁狂症（可空）
_FULL_BG_ROW = 77         # W77 完整背景故事（小说式段落）
_FULL_BG_COL = 23
# 背包：F(col=6)/N(col=14) 列
_INV_COL_LEFT = 6
_INV_COL_RIGHT = 14

# 字段标签关键词——读到这些说明单元格实际是标签而非数据
_LABEL_KEYWORDS = [
    "姓名", "玩家", "职业", "职业序号", "年龄", "性别", "住地", "故乡", "出生地", "时代",
    "力量", "敏捷", "意志", "体质", "外貌", "教育", "体型", "智力", "灵感", "幸运",
    "生命", "理智", "魔法", "移动",
]


def _is_label(text: str) -> bool:
    """判断文本是否像字段标签而非实际数据"""
    t = text.strip()
    if not t:
        return False
    # 含换行的多行文本一定是标签（如 "力量\nSTR"）
    if "\n" in t:
        return True
    for kw in _LABEL_KEYWORDS:
        if kw in t:
            return True
    return False


# ====================================================================
# 分块读取
# ====================================================================


def _parse_basic_info(ws) -> Tuple[str, str, int, str, str]:
    """提取姓名、性别、年龄、出生地、职业"""
    name = str(ws.cell(row=3, column=_COL_NAME).value or "").strip()
    if _is_label(name):
        name = ""
    # M6 是性别（本卡格式：M6="女"）
    gender_raw = ws.cell(row=6, column=_COL_NAME + 8).value  # M6
    gender = str(gender_raw or "").strip()
    if _is_label(gender):
        gender = ""
    # E6 是年龄
    age = _to_int(ws.cell(row=6, column=_COL_NAME).value)
    # E7 住地/出生地
    birthplace = str(ws.cell(row=7, column=_COL_NAME).value or "").strip()
    if _is_label(birthplace):
        birthplace = ""
    # E5 职业
    occ_name = str(ws.cell(row=5, column=_COL_NAME).value or "").strip()
    if _is_label(occ_name):
        occ_name = ""
    return name, gender, age, birthplace, occ_name


def _parse_stats(ws) -> Tuple[Dict[str, int], int]:
    """提取八项属性（STR..EDU 缩写键）与幸运值"""
    stats: Dict[str, int] = {}
    for label, col, row in _STAT_COLS:
        stats[label] = _to_int(ws.cell(row=row, column=col).value)
    luck = _to_int(ws.cell(row=_LUCK_ROW, column=_LUCK_COL).value)
    return stats, luck


def _parse_skills(ws) -> Dict[str, int]:
    """解析技能表，合并 init+growth+occ+interest 为最终值"""
    result: Dict[str, int] = {}

    def _read_side(side: dict) -> None:
        for r in range(_ROW_SKILL_START, _ROW_SKILL_END + 1):
            name_raw = ws.cell(row=r, column=side["name"]).value
            if not name_raw:
                continue
            name = _normalize(name_raw)
            if not name:
                continue
            init = _to_int(ws.cell(row=r, column=side["init"]).value)
            growth = _to_int(ws.cell(row=r, column=side["growth"]).value)
            occ = _to_int(ws.cell(row=r, column=side["occ"]).value)
            interest = _to_int(ws.cell(row=r, column=side["interest"]).value)
            total = init + growth + occ + interest
            if total > 0 or not result:  # 总是保留第一个技能确保占位
                result[name] = total

    _read_side(_SKILL_LEFT)
    _read_side(_SKILL_RIGHT)
    return result


def _parse_background(ws) -> Dict[str, str]:
    """提取七大背景项（AA 列，可跳过）"""
    bg: Dict[str, str] = {}
    for field, row in _BG_ROWS.items():
        raw = ws.cell(row=row, column=_BG_COL).value
        val = str(raw).strip() if raw else ""
        if val:
            bg[field] = val
    return bg


def _parse_phobias(ws) -> str:
    """提取恐惧症和躁狂症（AA75，部分角色卡可能为空）"""
    raw = ws.cell(row=_PHOBIA_ROW, column=_BG_COL).value
    return str(raw).strip() if raw else ""


def _parse_full_backstory(ws) -> str:
    """提取 W77 的小说式完整背景故事（非七项摘要）"""
    raw = ws.cell(row=_FULL_BG_ROW, column=_FULL_BG_COL).value
    return str(raw).strip() if raw else ""


def _parse_inventory(ws) -> List[str]:
    """提取背包物品（F 列和 N 列的自由文本）。

    单元格内换行是有意的卡格式（写卡时特意分行表示独立物品），整格作为一条保留，
    不要合并/拆分——真实卡实测（费利西蒂·马瑟斯.xlsx）即如此。
    """
    items: List[str] = []
    for r in range(_ROW_INV_START, _ROW_INV_END + 1):
        for col in (_INV_COL_LEFT, _INV_COL_RIGHT):
            raw = ws.cell(row=r, column=col).value
            val = str(raw).strip() if raw else ""
            if val and val not in ("状态", "部位", "物品名称", "背包格↓", ""):
                items.append(val)  # 状态：整格一条保留（换行是有意拆分）
    return items


# ====================================================================
# 衍生值
# ====================================================================


def _derive_hp_san_mp(stats: Dict[str, int]) -> Tuple[int, int, int]:
    """由八属性衍生 max HP/SAN/MP（对齐 glyphkeeper create_investigator）：
    max_hp = (CON + SIZ) // 2；max_mp = max(1, POW // 5)；max_san = POW。
    """
    con = stats.get("CON", 0)
    siz = stats.get("SIZ", 0)
    pow_ = stats.get("POW", 0)
    max_hp = (con + siz) // 2
    max_mp = max(1, pow_ // 5)
    max_san = pow_
    return max_hp, max_mp, max_san


# ====================================================================
# 主入口
# ====================================================================


def parse_investigator_xlsx(filepath: str | Path) -> Dict[str, Any]:
    """解析 xlsx 角色卡，返回 {meta, entity} 双层结构。

    - meta：读取到的原始元数据（name/gender/age/birthplace/occupation/luck），
      供展示与创建世界时选择角色；新架构无独立字段，不入库。
    - entity：可直接喂 storage.create_entity 的字段（entity_type/name/hp/mp/san/
      attributes_and_skills/inventory/background/tags），满状态入场。
    """
    import openpyxl  # 惰性导入：未装 openpyxl 时此处清晰报错

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["人物卡"]  # 状态：固定读"人物卡" sheet

    name, gender, age, birthplace, occupation = _parse_basic_info(ws)
    stats, luck = _parse_stats(ws)
    skills = _parse_skills(ws)
    background = _parse_background(ws)
    inventory = _parse_inventory(ws)
    full_backstory = _parse_full_backstory(ws)
    phobias = _parse_phobias(ws)
    if full_backstory:
        background["full_backstory"] = full_backstory
    if phobias:
        background["phobias_manias"] = phobias

    max_hp, max_mp, max_san = _derive_hp_san_mp(stats)
    attributes_and_skills = {**stats, **skills}

    entity = {
        "entity_type": "PC",
        "name": name,
        "occupation": occupation,
        "hp": max_hp,
        "hp_max": max_hp,
        "mp": max_mp,
        "mp_max": max_mp,
        "san": max_san,
        "san_max": max_san,
        "attributes_and_skills": attributes_and_skills,
        "inventory": [{"name": i} for i in inventory],
        "background": background,
        "tags": [],
    }

    meta = {
        "name": name,
        "gender": gender,
        "age": age,
        "birthplace": birthplace,
        "occupation": occupation,
        "luck": luck,
    }
    logger.info(
        "角色卡读取完成: %s (%s) STR=%d CON=%d ...",
        name, occupation, stats.get("STR", 0), stats.get("CON", 0),
    )
    return {"meta": meta, "entity": entity}
